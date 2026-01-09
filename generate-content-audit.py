#!/usr/bin/env python3
"""
Generate YWP-Content-Audit.xlsx
Creates a comprehensive audit of all YourWillPro content in Excel format.
"""

import os
import re
import json
from datetime import datetime
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

# Base directory
BASE_DIR = Path("/Users/donmacintosh/Desktop/yourwillpro-content")

def parse_frontmatter(content):
    """Extract frontmatter from markdown file."""
    match = re.match(r'^---\s*\n(.*?)\n---\s*\n', content, re.DOTALL)
    if not match:
        return {}

    frontmatter = {}
    lines = match.group(1).split('\n')
    current_key = None

    for line in lines:
        if ':' in line and not line.startswith(' '):
            key, value = line.split(':', 1)
            key = key.strip()
            value = value.strip().strip('"').strip("'")
            frontmatter[key] = value
            current_key = key
        elif current_key and line.strip():
            # Multi-line value
            frontmatter[current_key] += ' ' + line.strip()

    return frontmatter

def format_date(filepath):
    """Format file modification date as '9 Jan 2025 at 3:55 pm'"""
    try:
        timestamp = os.path.getmtime(filepath)
        dt = datetime.fromtimestamp(timestamp)

        # Get month abbreviation
        months = ['Jan', 'Feb', 'Mar', 'Apr', 'May', 'Jun',
                  'Jul', 'Aug', 'Sep', 'Oct', 'Nov', 'Dec']
        month = months[dt.month - 1]

        # Format time
        hour = dt.hour
        meridiem = 'am' if hour < 12 else 'pm'
        if hour == 0:
            hour = 12
        elif hour > 12:
            hour -= 12

        return f"{dt.day} {month} {dt.year} at {hour}:{dt.minute:02d} {meridiem}"
    except:
        return "Unknown"

def get_user_stories():
    """Scan user stories content."""
    stories_dir = BASE_DIR / "src/content/user-stories"
    stories = []

    for md_file in stories_dir.glob("*.md"):
        if md_file.name == "_section.md":
            continue

        try:
            content = md_file.read_text(encoding='utf-8')
            frontmatter = parse_frontmatter(content)

            title = frontmatter.get('title', md_file.stem)
            featured = frontmatter.get('featured', '').lower() == 'true'

            location = "User Stories - Featured story" if featured else "User Stories - Grid"
            description = frontmatter.get('summary', frontmatter.get('description', ''))

            stories.append({
                'location_on_website': location,
                'filename': md_file.name,
                'file_path': str(md_file.relative_to(BASE_DIR)),
                'description': description,
                'file_type': 'Content',
                'last_updated': format_date(md_file)
            })
        except Exception as e:
            print(f"Error processing {md_file}: {e}")

    return stories

def get_guides():
    """Scan guides content."""
    guides_dir = BASE_DIR / "src/content/guides"
    guides = []

    for md_file in guides_dir.glob("*.md"):
        try:
            content = md_file.read_text(encoding='utf-8')
            frontmatter = parse_frontmatter(content)

            title = frontmatter.get('title', md_file.stem)
            category = frontmatter.get('guideCategory', 'General')

            location = f"Guides - {category} section"
            description = frontmatter.get('summary', frontmatter.get('description', ''))

            guides.append({
                'location_on_website': location,
                'filename': md_file.name,
                'file_path': str(md_file.relative_to(BASE_DIR)),
                'description': description,
                'file_type': 'Content',
                'last_updated': format_date(md_file)
            })
        except Exception as e:
            print(f"Error processing {md_file}: {e}")

    return guides

def get_common_questions():
    """Scan common questions content."""
    questions_dir = BASE_DIR / "src/content/common-questions"
    questions = []

    for md_file in questions_dir.glob("*.md"):
        try:
            content = md_file.read_text(encoding='utf-8')
            frontmatter = parse_frontmatter(content)

            title = frontmatter.get('title', frontmatter.get('question', md_file.stem))

            questions.append({
                'location_on_website': 'Common Questions - FAQ list',
                'filename': md_file.name,
                'file_path': str(md_file.relative_to(BASE_DIR)),
                'description': title,
                'file_type': 'Content',
                'last_updated': format_date(md_file)
            })
        except Exception as e:
            print(f"Error processing {md_file}: {e}")

    return questions

def determine_image_usage(filename, filepath):
    """Determine where an image is used on the website."""
    filename_lower = filename.lower()
    path_parts = filepath.parts

    # Check by directory and filename patterns
    if 'logo' in filename_lower:
        return "Site branding"
    elif 'og-' in filename_lower or 'social' in filename_lower:
        return "Social sharing"
    elif 'ui' in path_parts:
        return "UI elements"
    elif 'hero' in filename_lower:
        # Try to determine which story/page
        if 'home' in filename_lower:
            return "Homepage hero"
        else:
            return "Hero image"
    elif 'user-stories' in path_parts:
        # Extract story name from path or filename
        story_name = filename_lower.replace('.png', '').replace('.jpg', '').replace('.jpeg', '').replace('-', ' ').title()
        return f"{story_name} story - hero image"
    elif 'guides' in path_parts:
        return "Guide content"
    else:
        return "Website content"

def get_images():
    """Scan all images."""
    images_dir = BASE_DIR / "public/images"
    images = []

    # Image extensions to include
    image_extensions = {'.png', '.jpg', '.jpeg', '.gif', '.svg', '.webp', '.ico'}

    for img_file in images_dir.rglob("*"):
        if img_file.is_file() and img_file.suffix.lower() in image_extensions:
            try:
                relative_path = img_file.relative_to(BASE_DIR)
                location = determine_image_usage(img_file.name, relative_path)

                images.append({
                    'location_on_website': location,
                    'filename': img_file.name,
                    'file_path': str(relative_path),
                    'description': f"Image: {img_file.stem.replace('-', ' ').replace('_', ' ').title()}",
                    'file_type': 'Image',
                    'last_updated': format_date(img_file)
                })
            except Exception as e:
                print(f"Error processing {img_file}: {e}")

    return images

def get_standalone_pages():
    """Scan standalone Astro pages."""
    pages_dir = BASE_DIR / "src/pages"
    pages = []

    # Pages to include (top-level only, no dynamic routes)
    include_pages = [
        'index.astro', 'about.astro', 'our-story.astro', 'feedback.astro',
        'start-here.astro', 'privacy.astro', 'terms.astro', 'disclaimer.astro',
        'sitemap.astro'
    ]

    # Mapping of filenames to descriptive locations
    page_locations = {
        'index.astro': 'Homepage',
        'about.astro': 'About Us page',
        'our-story.astro': 'Our Story page',
        'feedback.astro': 'Feedback page',
        'start-here.astro': 'Start Here page',
        'privacy.astro': 'Privacy Policy page',
        'terms.astro': 'Terms & Conditions page',
        'disclaimer.astro': 'Disclaimer page',
        'sitemap.astro': 'Sitemap page'
    }

    for page_name in include_pages:
        page_file = pages_dir / page_name
        if page_file.exists():
            try:
                location = page_locations.get(page_name, page_name.replace('.astro', '').replace('-', ' ').title())

                pages.append({
                    'location_on_website': location,
                    'filename': page_file.name,
                    'file_path': str(page_file.relative_to(BASE_DIR)),
                    'description': f"Standalone page: {location}",
                    'file_type': 'Page',
                    'last_updated': format_date(page_file)
                })
            except Exception as e:
                print(f"Error processing {page_file}: {e}")

    return pages

def create_excel_file(all_data, output_path):
    """Create the Excel file with proper formatting."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Content Audit"

    # Define headers
    headers = ['#', 'Location on website', 'Filename', 'Location', 'Description', 'File Type', 'Last Updated']

    # Write headers
    ws.append(headers)

    # Style the header row
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")

    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="left", vertical="center")

    # Add data rows
    row_num = 1
    for item in all_data:
        row_num += 1
        ws.append([
            row_num - 1,  # Row number (excluding header)
            item['location_on_website'],
            item['filename'],
            item['file_path'],
            item['description'],
            item['file_type'],
            item['last_updated']
        ])

    # Freeze the header row
    ws.freeze_panes = "A2"

    # Set column widths
    column_widths = {
        'A': 5,   # #
        'B': 40,  # Location on website
        'C': 30,  # Filename
        'D': 50,  # Location (file path)
        'E': 60,  # Description
        'F': 12,  # File Type
        'G': 25   # Last Updated
    }

    for col_letter, width in column_widths.items():
        ws.column_dimensions[col_letter].width = width

    # Save the workbook
    wb.save(output_path)

    return row_num - 1  # Total rows excluding header

def main():
    """Main execution function."""
    print("Generating YWP Content Audit...")
    print()

    # Collect all content
    print("Scanning user stories...")
    user_stories = get_user_stories()
    print(f"  Found {len(user_stories)} user stories")

    print("Scanning guides...")
    guides = get_guides()
    print(f"  Found {len(guides)} guides")

    print("Scanning common questions...")
    questions = get_common_questions()
    print(f"  Found {len(questions)} common questions")

    print("Scanning images...")
    images = get_images()
    print(f"  Found {len(images)} images")

    print("Scanning standalone pages...")
    pages = get_standalone_pages()
    print(f"  Found {len(pages)} standalone pages")

    print()

    # Combine all data
    all_data = user_stories + guides + questions + images + pages

    # Sort by File Type, then Location on website
    all_data.sort(key=lambda x: (x['file_type'], x['location_on_website']))

    # Determine output path
    output_paths = [
        Path("/mnt/user-data/outputs/YWP-Content-Audit.xlsx"),
        BASE_DIR / "outputs/YWP-Content-Audit.xlsx"
    ]

    output_path = None
    for path in output_paths:
        try:
            path.parent.mkdir(parents=True, exist_ok=True)
            output_path = path
            break
        except:
            continue

    if not output_path:
        print("ERROR: Could not create output directory")
        return

    # Create Excel file
    print(f"Creating Excel file at {output_path}...")
    total_rows = create_excel_file(all_data, output_path)

    # Calculate statistics
    content_count = len([x for x in all_data if x['file_type'] == 'Content'])
    image_count = len([x for x in all_data if x['file_type'] == 'Image'])
    page_count = len([x for x in all_data if x['file_type'] == 'Page'])

    # Get file size
    file_size = output_path.stat().st_size
    if file_size > 1024 * 1024:
        file_size_str = f"{file_size / (1024 * 1024):.2f} MB"
    else:
        file_size_str = f"{file_size / 1024:.2f} KB"

    # Print summary
    print()
    print("=" * 70)
    print("EXCEL FILE CREATED SUCCESSFULLY")
    print("=" * 70)
    print()
    print(f"Total rows: {total_rows} (excluding header)")
    print()
    print("Breakdown by File Type:")
    print(f"  - Content: {content_count} rows")
    print(f"  - Image:   {image_count} rows")
    print(f"  - Page:    {page_count} rows")
    print()
    print(f"File location: {output_path}")
    print(f"File size:     {file_size_str}")
    print()
    print("Ready for download!")
    print()

if __name__ == "__main__":
    main()
